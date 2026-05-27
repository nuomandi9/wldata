"""Integration tests for the fixed-report API.

Seeds an admin user, a tiny dictionary (vehicle/driver/route), and 3 delivery
records, then runs /api/report endpoints against them.

Also asserts that the parameter binding contract holds: SQL-injection-shaped
inputs land as literal parameter values (no rewriting of the SQL).
"""
import openpyxl
import pytest
import io
from sqlalchemy import select

from models.report import ReportTemplate


async def get_admin_token(client):
    resp = await client.post(
        "/api/auth/login", json={"username": "admin", "password": "admin123"}
    )
    return resp.json()["access_token"]


def auth_header(token):
    return {"Authorization": f"Bearer {token}"}


SAMPLE_TEMPLATE = {
    "key": "test_delivery_detail",
    "name": "测试明细",
    "description": "测试用报表",
    "sql_template": """
        SELECT
          dr.record_date AS record_date,
          v.plate_number AS plate_number,
          dro.route_code AS route_code,
          dr.customer_count AS customer_count,
          dr.delivery_count AS delivery_count
        FROM biz_delivery_record dr
        LEFT JOIN dict_vehicle v ON v.id = dr.vehicle_id
        LEFT JOIN dict_route dro ON dro.id = dr.route_id
        WHERE dr.record_date BETWEEN :start_date AND :end_date
          AND (CAST(:route_id AS INTEGER) IS NULL OR dr.route_id = CAST(:route_id AS INTEGER))
        ORDER BY dr.record_date DESC
    """,
    "params_schema": [
        {"name": "start_date", "label": "开始日期", "type": "date", "required": True},
        {"name": "end_date", "label": "结束日期", "type": "date", "required": True},
        {"name": "route_id", "label": "线路", "type": "int", "required": False},
    ],
    "columns_schema": [
        {"key": "record_date", "label": "日期", "type": "date"},
        {"key": "plate_number", "label": "车牌号", "type": "str"},
        {"key": "route_code", "label": "线路编码", "type": "str"},
        {"key": "customer_count", "label": "客户数", "type": "int"},
        {"key": "delivery_count", "label": "送货数", "type": "int"},
    ],
}


async def seed_test_data(client, headers):
    """Three delivery records: 2 on route R001, 1 on route R002."""
    p = await client.post(
        "/api/dict/person",
        json={"name": "张司机", "id_card": "110101199001011234"},
        headers=headers,
    )
    person_id = p.json()["id"]

    v = await client.post(
        "/api/dict/vehicle",
        json={"plate_number": "京A12345", "driver_id": person_id},
        headers=headers,
    )
    vehicle_id = v.json()["id"]

    r1 = await client.post(
        "/api/dict/route",
        json={"route_code": "R001", "route_name": "城东一线", "driver_id": person_id},
        headers=headers,
    )
    route_id_1 = r1.json()["id"]

    r2 = await client.post(
        "/api/dict/route",
        json={"route_code": "R002", "route_name": "城西二线", "driver_id": person_id},
        headers=headers,
    )
    route_id_2 = r2.json()["id"]

    # Insert delivery records via the import endpoint to avoid bypassing the model
    headers_list = ["日期", "车牌号", "驾驶员", "线路编码", "客户数", "送货数"]
    rows = [
        ["2026-05-01", "京A12345", "张司机", "R001", 10, 25],
        ["2026-05-02", "京A12345", "张司机", "R001", 12, 30],
        ["2026-05-03", "京A12345", "张司机", "R002", 8, 18],
    ]
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(headers_list)
    for row in rows:
        ws.append(row)
    bio = io.BytesIO()
    wb.save(bio)

    files = {"file": ("data.xlsx", bio.getvalue(),
                      "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    preview = await client.post(
        "/api/import/delivery_record/preview", files=files, headers=headers,
    )
    assert preview.status_code == 200, preview.text
    row_numbers = [r["row_number"] for r in preview.json()["rows"]]

    bio.seek(0)
    files = {"file": ("data.xlsx", bio.getvalue(),
                      "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    commit = await client.post(
        f"/api/import/delivery_record/commit",
        files=files,
        data={
            "commit_row_numbers": str(row_numbers).replace("'", '"'),
            "warn_notes": "{}",
        },
        headers=headers,
    )
    assert commit.status_code == 200, commit.text
    return {"route_id_1": route_id_1, "route_id_2": route_id_2}


async def seed_template():
    """Insert one report template directly into the test DB."""
    from tests.conftest import test_session
    async with test_session() as session:
        session.add(ReportTemplate(**SAMPLE_TEMPLATE))
        await session.commit()


@pytest.mark.asyncio
async def test_list_templates_requires_auth(client, seed_admin):
    resp = await client.get("/api/report/templates")
    assert resp.status_code == 403


@pytest.mark.asyncio
async def test_list_templates(client, seed_admin):
    await seed_template()
    token = await get_admin_token(client)
    resp = await client.get("/api/report/templates", headers=auth_header(token))
    assert resp.status_code == 200
    body = resp.json()
    assert len(body) == 1
    assert body[0]["key"] == "test_delivery_detail"
    assert body[0]["params_schema"][0]["name"] == "start_date"
    assert "sql_template" not in body[0]  # SQL is never leaked to clients


@pytest.mark.asyncio
async def test_execute_with_date_range(client, seed_admin):
    await seed_template()
    token = await get_admin_token(client)
    headers = auth_header(token)
    await seed_test_data(client, headers)

    resp = await client.post(
        "/api/report/test_delivery_detail/execute",
        json={"params": {"start_date": "2026-05-01", "end_date": "2026-05-31"}},
        headers=headers,
    )
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert body["total"] == 3
    assert len(body["rows"]) == 3
    assert {r["route_code"] for r in body["rows"]} == {"R001", "R002"}


@pytest.mark.asyncio
async def test_execute_with_optional_route_filter(client, seed_admin):
    await seed_template()
    token = await get_admin_token(client)
    headers = auth_header(token)
    ids = await seed_test_data(client, headers)

    resp = await client.post(
        "/api/report/test_delivery_detail/execute",
        json={"params": {
            "start_date": "2026-05-01",
            "end_date": "2026-05-31",
            "route_id": ids["route_id_1"],
        }},
        headers=headers,
    )
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert body["total"] == 2
    assert all(r["route_code"] == "R001" for r in body["rows"])


@pytest.mark.asyncio
async def test_execute_missing_required_param(client, seed_admin):
    await seed_template()
    token = await get_admin_token(client)
    headers = auth_header(token)

    resp = await client.post(
        "/api/report/test_delivery_detail/execute",
        json={"params": {"start_date": "2026-05-01"}},  # missing end_date
        headers=headers,
    )
    assert resp.status_code == 400
    assert "结束日期" in resp.json()["detail"]


@pytest.mark.asyncio
async def test_execute_pagination(client, seed_admin):
    await seed_template()
    token = await get_admin_token(client)
    headers = auth_header(token)
    await seed_test_data(client, headers)

    resp = await client.post(
        "/api/report/test_delivery_detail/execute",
        json={
            "params": {"start_date": "2026-05-01", "end_date": "2026-05-31"},
            "page": 1, "page_size": 2,
        },
        headers=headers,
    )
    body = resp.json()
    assert body["total"] == 3
    assert len(body["rows"]) == 2


@pytest.mark.asyncio
async def test_export_returns_xlsx(client, seed_admin):
    await seed_template()
    token = await get_admin_token(client)
    headers = auth_header(token)
    await seed_test_data(client, headers)

    resp = await client.post(
        "/api/report/test_delivery_detail/export",
        json={"params": {"start_date": "2026-05-01", "end_date": "2026-05-31"}},
        headers=headers,
    )
    assert resp.status_code == 200
    assert resp.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    # xlsx is a zip file — magic bytes PK\x03\x04
    assert resp.content[:4] == b"PK\x03\x04"

    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    ws = wb.active
    # Header row + 3 data rows
    assert ws.max_row == 4
    header_values = [c.value for c in ws[1]]
    assert "日期" in header_values
    assert "车牌号" in header_values


@pytest.mark.asyncio
async def test_execute_unknown_template_404(client, seed_admin):
    token = await get_admin_token(client)
    resp = await client.post(
        "/api/report/does_not_exist/execute",
        json={"params": {}},
        headers=auth_header(token),
    )
    assert resp.status_code == 404


@pytest.mark.asyncio
async def test_param_with_quote_does_not_break_sql(client, seed_admin):
    """Whatever the user puts in a param value stays a bound value — it cannot
    rewrite the SQL. Here we pass a string with a quote in place of an int —
    the executor should reject it with a 400 (param type error), NOT execute
    arbitrary SQL or 500 with a syntax error.
    """
    await seed_template()
    token = await get_admin_token(client)
    headers = auth_header(token)

    resp = await client.post(
        "/api/report/test_delivery_detail/execute",
        json={"params": {
            "start_date": "2026-05-01",
            "end_date": "2026-05-31",
            "route_id": "1; DROP TABLE biz_delivery_record; --",
        }},
        headers=headers,
    )
    assert resp.status_code == 400
    assert "线路" in resp.json()["detail"]
