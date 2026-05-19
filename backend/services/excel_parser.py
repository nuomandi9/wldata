"""Parse uploaded Excel workbooks into raw row dicts according to a template.

This stage is lossy on purpose — values are extracted but **not** validated
or coerced into FK ids; that work happens in `import_validator`. Whitespace
is trimmed and empty cells become None.
"""
from datetime import date, datetime
from io import BytesIO
from typing import Any

from openpyxl import load_workbook

from services.import_templates import ImportTemplate


class ExcelParseError(Exception):
    """Raised when the uploaded file is not a valid xlsx for the template."""


def parse_workbook(file_bytes: bytes, template: ImportTemplate) -> list[dict[str, Any]]:
    """Read the first sheet of an xlsx, map headers to template fields, return raw rows.

    The first row is treated as the header row; columns not declared on the
    template are ignored. Cells are normalized (strip strings, empty → None)
    but **not** type-coerced — the validator catches type mismatches and
    reports them as BLOCK issues with the row number.
    """
    try:
        wb = load_workbook(BytesIO(file_bytes), data_only=True, read_only=True)
    except Exception as exc:
        raise ExcelParseError(f"无法读取 Excel 文件：{exc}") from exc

    ws = wb.active
    if ws is None or ws.max_row < 2:
        raise ExcelParseError("Excel 文件为空或缺少数据行")

    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if header_row is None:
        raise ExcelParseError("Excel 文件没有表头")

    header_to_col = {col.header: col for col in template.columns}
    declared = set(header_to_col)
    sheet_headers = [_normalize_cell(h) for h in header_row]
    sheet_header_set = {h for h in sheet_headers if h}

    missing_required = [
        col.header for col in template.columns
        if col.required and col.header not in sheet_header_set
    ]
    if missing_required:
        raise ExcelParseError(f"Excel 表头缺少必填列：{', '.join(missing_required)}")

    # Map (column index in sheet → ImportColumn)
    column_map: dict[int, Any] = {}
    for idx, header in enumerate(sheet_headers):
        if header in declared:
            column_map[idx] = header_to_col[header]

    rows: list[dict[str, Any]] = []
    for row in rows_iter:
        if row is None or all(_is_empty(cell) for cell in row):
            continue
        record: dict[str, Any] = {}
        for idx, col in column_map.items():
            value = row[idx] if idx < len(row) else None
            record[col.field] = _normalize_cell(value)
        rows.append(record)
    return rows


def _normalize_cell(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        v = value.strip()
        return v or None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return value


def _is_empty(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str) and not value.strip():
        return True
    return False
