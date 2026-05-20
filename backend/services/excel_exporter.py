"""Render report execution results to xlsx bytes.

Uses standard openpyxl Workbook (not write_only) — simpler, and adequate
for the row counts this app produces. If exports ever exceed ~50k rows
we can switch to write_only with WriteOnlyCell for memory savings.
"""
from datetime import date, datetime
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


HEADER_FILL = PatternFill("solid", fgColor="0A1628")
HEADER_FONT = Font(name="Microsoft YaHei", bold=True, color="F5F0EB")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")


def render_xlsx(
    sheet_name: str,
    columns: list[dict],
    rows: list[dict[str, Any]],
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = _safe_sheet_name(sheet_name)

    for col_idx, col in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col["label"])
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, col in enumerate(columns, start=1):
            ws.cell(
                row=row_idx,
                column=col_idx,
                value=_cell_value(row.get(col["key"]), col.get("type", "str")),
            )

    for col_idx, col in enumerate(columns, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max(
            12, min(40, len(col["label"]) * 2 + 4)
        )

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def _cell_value(value: Any, typ: str) -> Any:
    if value is None:
        return None
    if typ == "date" and isinstance(value, str):
        try:
            return date.fromisoformat(value)
        except ValueError:
            return value
    if typ == "datetime" and isinstance(value, str):
        try:
            return datetime.fromisoformat(value)
        except ValueError:
            return value
    return value


def _safe_sheet_name(name: str) -> str:
    """Excel sheet names are limited to 31 chars and disallow []:*?/\\."""
    cleaned = "".join("_" if ch in "[]:*?/\\" else ch for ch in name)
    return cleaned[:31] or "Sheet1"
